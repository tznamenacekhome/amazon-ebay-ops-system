"""Manually import non-eBay purchase COGS source rows from a Google Sheets XLSX export.

These rows are source data for sales-order COGS matching. They intentionally
live outside the eBay purchases workflow and never write purchases or
purchase_items.

Do not add this importer to run_all_syncs.py or scheduled data-sync jobs. The
long-term workflow for this source is MBOP -> TIM Sheet export/update, not
TIM Sheet -> MBOP scheduled sync.
"""

from __future__ import annotations

import argparse
import logging
import os
from dataclasses import dataclass
from datetime import UTC, date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client

LOGGER = logging.getLogger("import_non_ebay_cogs_sources")

DEFAULT_INPUT = Path("data/tim_non_ebay_purchases.xlsx")
DEFAULT_SOURCE_DOCUMENT_ID = "1-5bKK_gEpzl2lvj7dm26omc-Kof4efQenc69sQUe__I"
DEFAULT_SOURCE_DOCUMENT_TITLE = "TIM"
DEFAULT_FULFILLMENT_CHANNEL = "Prep-Center"
DEFAULT_CUTOFF_DATE = date(2025, 1, 1)
BATCH_SIZE = 250

SHEET_2024_COLUMN_MAP = {
    "order_date": 0,
    "supplier": 1,
    "asin": 2,
    "supplier_order_number": 3,
    "msku": 4,
    "description": 5,
    "size_color": 6,
    "bundles": 7,
    "quantity": 8,
    "received_by_prep_center_quantity": 9,
    "damaged_quantity": 10,
    "unit_cost": 11,
    "list_price": 12,
    "notes": 13,
    "expiration_date": 14,
    "tracking": 15,
    "remarks": 16,
}

SHEET_2023_COLUMN_MAP = {
    "order_date": 0,
    "supplier": 1,
    "asin": 2,
    "supplier_order_number": 3,
    "msku": 4,
    "description": 5,
    "size_color": 6,
    "quantity": 7,
    "received_by_prep_center_quantity": 8,
    "damaged_quantity": 9,
    "unit_cost": 10,
    "list_price": 11,
    "notes": 12,
    "expiration_date": 13,
    "tracking": 14,
    "remarks": 15,
}

COMPACT_PURCHASE_COLUMN_MAP = {
    "order_date": 0,
    "supplier": 1,
    "asin": 2,
    "supplier_order_number": 3,
    "description": 4,
    "quantity": 5,
    "unit_cost": 6,
}


@dataclass
class ImportRow:
    sheet_name: str
    row_number: int
    raw_values: list[Any]
    order_date: date
    asin: str
    supplier: str | None
    supplier_order_number: str | None
    msku: str | None
    description: str | None
    size_color: str | None
    bundles: int | None
    quantity: int | None
    received_by_prep_center_quantity: int | None
    damaged_quantity: int | None
    unit_cost: Decimal | None
    list_price: Decimal | None
    notes: str | None
    expiration_date: date | None
    tracking: str | None
    remarks: str | None


def main() -> int:
    args = parse_args()
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s - %(message)s",
    )
    logging.getLogger("httpx").setLevel(logging.WARNING)
    load_dotenv()

    try:
        rows = read_workbook(args.input, args.cutoff_date, args.quantity_source)
        print_summary(rows, args.cutoff_date)

        if not args.apply:
            LOGGER.info("Dry run complete. No Supabase writes performed.")
            return 0

        payload = [
            build_payload_row(
                row=row,
                source_document_id=args.source_document_id,
                source_document_title=args.source_document_title,
                fulfillment_channel=args.fulfillment_channel,
            )
            for row in rows
        ]

        supabase = get_supabase_client()
        for chunk in chunks(payload, BATCH_SIZE):
            supabase.table("non_ebay_purchase_cogs_sources").upsert(
                chunk,
                on_conflict="source_document_id,source_sheet_name,source_row_number",
            ).execute()

        LOGGER.info("Non-eBay COGS source rows upserted: %s", len(payload))
        return 0
    except Exception as error:  # noqa: BLE001 - integration should fail safely
        LOGGER.exception("Non-eBay COGS source import failed safely: %s", error)
        return 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Analyze or import non-eBay purchase COGS source rows."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="XLSX export of the Google Sheet.",
    )
    parser.add_argument(
        "--cutoff-date",
        type=parse_iso_date,
        default=DEFAULT_CUTOFF_DATE,
        help="Import rows with order_date on or after this date.",
    )
    parser.add_argument(
        "--source-document-id",
        default=DEFAULT_SOURCE_DOCUMENT_ID,
        help="Google Sheets document ID used for idempotent source tracking.",
    )
    parser.add_argument(
        "--source-document-title",
        default=DEFAULT_SOURCE_DOCUMENT_TITLE,
        help="Human-readable source document title.",
    )
    parser.add_argument(
        "--fulfillment-channel",
        default=DEFAULT_FULFILLMENT_CHANNEL,
        help="Fulfillment channel to assign to imported rows.",
    )
    parser.add_argument(
        "--quantity-source",
        choices=("ordered", "received"),
        default="received",
        help=(
            "For sheets with a received/prep-center quantity column, choose whether "
            "the stored quantity field should use ordered quantity or received quantity."
        ),
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Write rows to non_ebay_purchase_cogs_sources.",
    )
    return parser.parse_args()


def get_supabase_client():
    supabase_url = os.getenv("SUPABASE_URL")
    supabase_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")

    if not supabase_url or not supabase_key:
        raise RuntimeError(
            "Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY environment variable."
        )

    return create_client(supabase_url, supabase_key)


def read_workbook(
    path: Path,
    cutoff_date: date,
    quantity_source: str,
) -> list[ImportRow]:
    if not path.exists():
        raise FileNotFoundError(path)

    workbook = load_workbook(path, data_only=True, read_only=True)
    rows: list[ImportRow] = []

    for worksheet in workbook.worksheets:
        if worksheet.title == "2024":
            rows.extend(
                read_sheet(
                    worksheet=worksheet,
                    column_map=SHEET_2024_COLUMN_MAP,
                    first_data_row=3,
                    cutoff_date=cutoff_date,
                    quantity_source=quantity_source,
                )
            )
        elif worksheet.title == "2023":
            rows.extend(
                read_sheet(
                    worksheet=worksheet,
                    column_map=SHEET_2023_COLUMN_MAP,
                    first_data_row=1,
                    cutoff_date=cutoff_date,
                    quantity_source=quantity_source,
                )
            )
        else:
            detected_map = detect_header_map(worksheet)
            if detected_map:
                rows.extend(
                    read_sheet(
                        worksheet=worksheet,
                        column_map=detected_map,
                        first_data_row=2,
                        cutoff_date=cutoff_date,
                        quantity_source=quantity_source,
                    )
                )
            else:
                LOGGER.info("Skipping unsupported or blank sheet: %s", worksheet.title)

    return sorted(rows, key=lambda row: (row.order_date, row.sheet_name, row.row_number))


def read_sheet(
    *,
    worksheet,
    column_map: dict[str, int],
    first_data_row: int,
    cutoff_date: date,
    quantity_source: str,
) -> list[ImportRow]:
    rows: list[ImportRow] = []

    for row_number, values_tuple in enumerate(
        worksheet.iter_rows(min_row=first_data_row, values_only=True),
        start=first_data_row,
    ):
        raw_values = list(values_tuple)
        if is_blank_row(raw_values):
            continue

        order_date = to_date(get_cell(raw_values, column_map, "order_date"))
        if not order_date or order_date < cutoff_date:
            continue

        asin = clean_asin(get_cell(raw_values, column_map, "asin"))
        if not asin:
            LOGGER.warning(
                "Skipping %s row %s because ASIN is blank.",
                worksheet.title,
                row_number,
            )
            continue

        ordered_quantity = to_int(get_cell(raw_values, column_map, "quantity"))
        received_quantity = to_int(
            get_cell(raw_values, column_map, "received_by_prep_center_quantity")
        )
        stored_quantity = ordered_quantity
        if quantity_source == "received" and received_quantity is not None:
            stored_quantity = received_quantity

        rows.append(
            ImportRow(
                sheet_name=worksheet.title,
                row_number=row_number,
                raw_values=raw_values,
                order_date=order_date,
                asin=asin,
                supplier=to_text(get_cell(raw_values, column_map, "supplier")),
                supplier_order_number=to_identifier(
                    get_cell(raw_values, column_map, "supplier_order_number")
                ),
                msku=to_identifier(get_cell(raw_values, column_map, "msku")),
                description=to_text(get_cell(raw_values, column_map, "description")),
                size_color=to_text(get_cell(raw_values, column_map, "size_color")),
                bundles=to_int(get_cell(raw_values, column_map, "bundles")),
                quantity=stored_quantity,
                received_by_prep_center_quantity=received_quantity,
                damaged_quantity=to_int(
                    get_cell(raw_values, column_map, "damaged_quantity")
                ),
                unit_cost=to_decimal(get_cell(raw_values, column_map, "unit_cost")),
                list_price=to_decimal(get_cell(raw_values, column_map, "list_price")),
                notes=to_text(get_cell(raw_values, column_map, "notes")),
                expiration_date=to_date(
                    get_cell(raw_values, column_map, "expiration_date")
                ),
                tracking=to_identifier(get_cell(raw_values, column_map, "tracking")),
                remarks=to_text(get_cell(raw_values, column_map, "remarks")),
            )
        )

    return rows


def detect_header_map(worksheet) -> dict[str, int] | None:
    header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        return None

    normalized = {normalize_header(value): index for index, value in enumerate(header)}
    required = {"order_date", "supplier", "asin", "supplier_order_number", "description", "quantity", "unit_cost"}
    if not required.issubset(normalized):
        return None

    return {
        "order_date": normalized["order_date"],
        "supplier": normalized["supplier"],
        "asin": normalized["asin"],
        "supplier_order_number": normalized["supplier_order_number"],
        "description": normalized["description"],
        "quantity": normalized["quantity"],
        "unit_cost": normalized["unit_cost"],
    }


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    aliases = {
        "order date": "order_date",
        "supplier": "supplier",
        "asin": "asin",
        "order #": "supplier_order_number",
        "order number": "supplier_order_number",
        "description": "description",
        "qty": "quantity",
        "quantity": "quantity",
        "cost price": "unit_cost",
        "unit cost": "unit_cost",
    }
    return aliases.get(text, text.replace(" ", "_"))


def build_payload_row(
    *,
    row: ImportRow,
    source_document_id: str,
    source_document_title: str,
    fulfillment_channel: str,
) -> dict[str, Any]:
    return {
        "source_system": "google_sheets",
        "source_document_id": source_document_id,
        "source_document_title": source_document_title,
        "source_sheet_name": row.sheet_name,
        "source_row_number": row.row_number,
        "fulfillment_channel": fulfillment_channel,
        "order_date": row.order_date.isoformat(),
        "supplier": row.supplier,
        "asin": row.asin,
        "supplier_order_number": row.supplier_order_number,
        "msku": row.msku,
        "description": row.description,
        "size_color": row.size_color,
        "bundles": row.bundles,
        "quantity": row.quantity,
        "received_by_prep_center_quantity": row.received_by_prep_center_quantity,
        "damaged_quantity": row.damaged_quantity,
        "unit_cost": str(row.unit_cost) if row.unit_cost is not None else None,
        "list_price": str(row.list_price) if row.list_price is not None else None,
        "notes": row.notes,
        "expiration_date": row.expiration_date.isoformat()
        if row.expiration_date
        else None,
        "tracking": row.tracking,
        "remarks": row.remarks,
        "raw_row_json": {
            "values": [to_json_value(value) for value in row.raw_values],
        },
        "updated_at": datetime.now(UTC).isoformat(timespec="seconds"),
    }


def print_summary(rows: list[ImportRow], cutoff_date: date) -> None:
    if rows:
        min_date = min(row.order_date for row in rows)
        max_date = max(row.order_date for row in rows)
    else:
        min_date = None
        max_date = None

    by_sheet: dict[str, int] = {}
    units = 0
    extended_cost = Decimal("0")
    for row in rows:
        by_sheet[row.sheet_name] = by_sheet.get(row.sheet_name, 0) + 1
        units += row.quantity or 0
        if row.unit_cost is not None and row.quantity is not None:
            extended_cost += row.unit_cost * Decimal(row.quantity)

    LOGGER.info("Cutoff date: %s", cutoff_date.isoformat())
    LOGGER.info("Rows selected: %s", len(rows))
    LOGGER.info("Rows by sheet: %s", by_sheet)
    LOGGER.info("Order date range: %s to %s", min_date, max_date)
    LOGGER.info("Quantity total: %s", units)
    LOGGER.info("Extended unit-cost total: %s", round(extended_cost, 2))


def get_cell(values: list[Any], column_map: dict[str, int], key: str) -> Any:
    index = column_map.get(key)
    if index is None or index >= len(values):
        return None
    return values[index]


def is_blank_row(values: list[Any]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def clean_asin(value: Any) -> str | None:
    text = to_identifier(value)
    if not text:
        return None
    return text.upper()


def to_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def to_identifier(value: Any) -> str | None:
    text = to_text(value)
    if not text:
        return None
    return text.replace("\r\n", "\n").strip()


def to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(Decimal(str(value)))
    except (InvalidOperation, ValueError):
        LOGGER.warning("Unable to parse integer value: %r", value)
        return None


def to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value)).quantize(Decimal("0.0001"))
    except (InvalidOperation, ValueError):
        LOGGER.warning("Unable to parse decimal value: %r", value)
        return None


def to_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return datetime.combine(date(1899, 12, 30), time()).date() + timedelta(days=int(value))
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    LOGGER.warning("Unable to parse date value: %r", value)
    return None


def to_json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def parse_iso_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def chunks(rows: list[dict[str, Any]], size: int):
    for index in range(0, len(rows), size):
        yield rows[index : index + size]


if __name__ == "__main__":
    raise SystemExit(main())
